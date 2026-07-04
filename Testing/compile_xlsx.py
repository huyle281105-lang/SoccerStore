import os
import csv
import glob
import copy
import openpyxl

def update_xlsx():
    xlsx_file = os.path.join(os.path.dirname(__file__), "SoccerStore_TestCases.xlsx")
    csv_dir = os.path.join(os.path.dirname(__file__), "TestCases")
    
    if not os.path.exists(xlsx_file):
        print(f"Error: {xlsx_file} not found!")
        return

    print("Loading original Excel workbook...")
    wb = openpyxl.load_workbook(xlsx_file)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        # Find the matching CSV file
        csv_pattern = os.path.join(csv_dir, f"{sheet_name}*.csv")
        csv_files = glob.glob(csv_pattern)
        
        if not csv_files:
            print(f"Warning: No matching CSV file found for sheet '{sheet_name}'")
            continue
            
        csv_file = csv_files[0]
        print(f"Updating sheet '{sheet_name}' from {os.path.basename(csv_file)}...")
        
        ws = wb[sheet_name]

        # --- Step 1: Detect the standard test case style/height from TC03+ rows ---
        # Find the first header row (containing "Test Case#") to know where TCs start
        tc_header_row = None
        for r in range(1, 15):
            if ws.cell(r, 1).value == 'Test Case#':
                tc_header_row = r
                break
        
        if tc_header_row is None:
            print(f"  Could not find 'Test Case#' header row in sheet '{sheet_name}', skipping style fix.")
        else:
            # TC rows start at tc_header_row+1
            tc_start = tc_header_row + 1
            
            # The standard style comes from TC03 (third TC row = tc_start+2)
            ref_row = tc_start + 2
            ref_height = ws.row_dimensions[ref_row].height or 65.0

            # Fix TC01 (tc_start) and TC02 (tc_start+1) row heights and cell styles
            for fix_row in [tc_start, tc_start + 1]:
                # Set row height to match standard TC rows
                ws.row_dimensions[fix_row].height = ref_height
                
                # Copy cell styles from the reference row (TC03)
                for col in range(1, ws.max_column + 1):
                    src_cell = ws.cell(row=ref_row, column=col)
                    dst_cell = ws.cell(row=fix_row, column=col)
                    if dst_cell.__class__.__name__ == 'Cell' and src_cell.__class__.__name__ == 'Cell':
                        dst_cell.font = copy.copy(src_cell.font)
                        dst_cell.fill = copy.copy(src_cell.fill)
                        dst_cell.border = copy.copy(src_cell.border)
                        dst_cell.alignment = copy.copy(src_cell.alignment)
                        dst_cell.number_format = src_cell.number_format
            
            print(f"  Fixed TC01/TC02 row height to {ref_height}pt and synchronized styles.")

        # --- Step 2: Update cell values from CSV ---
        max_csv_row = 0
        with open(csv_file, mode='r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                max_csv_row = max(max_csv_row, row_idx)
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.__class__.__name__ == 'Cell':
                        cell.value = value
        
        # --- Step 3: Clean up any leftover duplicate/extra TC rows beyond the CSV ---
        for r in range(max_csv_row + 1, ws.max_row + 1):
            first_cell = ws.cell(r, 1)
            if first_cell.__class__.__name__ == 'Cell' and first_cell.value and str(first_cell.value).startswith('TC'):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.__class__.__name__ == 'Cell':
                        cell.value = None

    wb.save(xlsx_file)
    print(f"\nDone! Successfully updated {xlsx_file}")
    print("TC01 and TC02 now have the same row height and style as all other test case rows.")

if __name__ == "__main__":
    update_xlsx()
